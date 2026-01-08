#!/usr/bin/env python3
"""
对账神器 - 批量处理发票/收据并生成Excel对账单
支持：中文、日文、英文等多语言收据
支持：PDF、图片、单收据、多收据
使用OpenAI GPT-4o Vision API
"""
import os
import pandas as pd
from tqdm import tqdm
from dotenv import load_dotenv
from openai_vision_extractor import OpenAIVisionExtractor
from pdf2image import convert_from_path
import tempfile

# 加载环境变量
load_dotenv()

# --- 配置区 ---
INPUT_FOLDER = "/Users/esther/Downloads/consolidated_receipts"  # 输入文件夹
OUTPUT_EXCEL = "我的对账单.xlsx"  # 输出Excel文件名
OPENAI_VISION_API_KEY = os.getenv("OPENAI_VISION_API_KEY")  # OpenAI Vision API Key


def process_file(file_path: str, extractor: OpenAIVisionExtractor) -> list:
    """
    处理单个文件（支持PDF、图片、单收据、多收据）

    :param file_path: 文件路径
    :param extractor: OpenAI Vision提取器
    :return: 收据列表（支持多个收据）
    """
    ext = os.path.splitext(file_path)[1].lower()
    filename = os.path.basename(file_path)

    try:
        # PDF处理：转换为图片后识别
        if ext == '.pdf':
            print(f"  📄 PDF文件，转换为图片...")
            with tempfile.TemporaryDirectory() as temp_dir:
                # 转换PDF为图片（提高DPI以获得更清晰的识别）
                images = convert_from_path(file_path, dpi=300)

                all_receipts = []
                for page_num, image in enumerate(images, 1):
                    # 保存为临时文件
                    temp_image_path = os.path.join(temp_dir, f"page_{page_num}.jpg")
                    image.save(temp_image_path, 'JPEG')

                    # 识别这一页
                    print(f"    📖 第{page_num}页识别中...")
                    receipts = extractor.extract_from_image(temp_image_path)

                    # 为每个收据添加源文件信息
                    for receipt in receipts:
                        receipt['源文件名'] = f"{filename} (第{page_num}页)"

                    all_receipts.extend(receipts)

                print(f"  ✅ PDF识别完成：{len(all_receipts)}个收据")
                return all_receipts

        # 图片处理：直接识别
        elif ext in ['.jpg', '.png', '.jpeg', '.bmp']:
            print(f"  🖼️ 图片文件识别中...")
            receipts = extractor.extract_from_image(file_path)

            # 为每个收据添加源文件信息
            for receipt in receipts:
                receipt['源文件名'] = filename

            if len(receipts) > 1:
                print(f"  ✅ 识别到{len(receipts)}个收据")
            else:
                print(f"  ✅ 提取成功")

            return receipts

        else:
            print(f"  ⚠️ 不支持的文件格式: {ext}")
            return [{"源文件名": filename, "错误": f"不支持的文件格式: {ext}"}]

    except Exception as e:
        print(f"  ❌ 处理失败: {e}")
        import traceback
        traceback.print_exc()
        return [{"源文件名": filename, "错误": str(e)}]


def convert_receipt_to_row(receipt: dict) -> dict:
    """
    将收据字典转换为Excel行格式

    :param receipt: 收据信息字典
    :return: Excel行字典
    """
    return {
        "日期": receipt.get("issue_date"),
        "店铺/公司名称": receipt.get("seller_name"),
        "价税合计": receipt.get("total_amount"),
        "货币": receipt.get("currency"),
        "源文件名": receipt.get("源文件名"),
        "商品列表": receipt.get("items"),
    }


def main():
    """主处理流程"""
    if not OPENAI_VISION_API_KEY:
        print("❌ 请设置环境变量: OPENAI_VISION_API_KEY")
        print("   获取方式: https://platform.openai.com/api-keys")
        return

    # 初始化 OpenAI Vision
    extractor = OpenAIVisionExtractor(OPENAI_VISION_API_KEY)

    # 获取所有支持的文件（图片 + PDF）
    supported_extensions = ('.jpg', '.png', '.jpeg', '.bmp', '.pdf')
    files = [f for f in os.listdir(INPUT_FOLDER)
             if f.lower().endswith(supported_extensions)]

    # 过滤掉保险单（没有实际支付金额）
    files = [f for f in files if not f.lower().startswith('insurance')]

    if not files:
        print(f"❌ 在 {INPUT_FOLDER} 中没有找到支持的文件")
        return

    # 小批量处理：只处理前20个文件
    BATCH_MODE = True
    BATCH_SIZE = 20
    # 只处理PDF文件
    PDF_ONLY = True  # 只处理PDF

    if BATCH_MODE:
        if PDF_ONLY:
            files = [f for f in files if f.lower().endswith('.pdf')][:BATCH_SIZE]
            if not files:
                print("❌ 没有找到PDF文件")
                return
        else:
            files = files[:BATCH_SIZE]
        print(f"📦 批量模式：只处理前 {len(files)} 个文件\n")
    else:
        # 只处理PDF文件
        if PDF_ONLY:
            files = [f for f in files if f.lower().endswith('.pdf')]
            print(f"📂 找到 {len(files)} 个PDF文件\n")
        else:
            print(f"📂 找到 {len(files)} 个文件\n")
    print("="*60)

    all_results = []

    # 遍历处理每个文件
    for i, filename in enumerate(files, 1):
        file_path = os.path.join(INPUT_FOLDER, filename)
        print(f"\n[{i}/{len(files)}] 处理: {filename}")

        # 处理文件（支持PDF、多收据）
        receipts = process_file(file_path, extractor)

        # 转换为Excel行格式
        for receipt in receipts:
            row = convert_receipt_to_row(receipt)
            all_results.append(row)

    # 生成 Excel
    if all_results:
        print("\n" + "="*60)
        print(f"📊 成功处理 {len(all_results)} 条收据记录，正在生成Excel...")

        df = pd.DataFrame(all_results)

        # 调整列顺序（只保留需要的列）
        priority_cols = [
            "日期", "店铺/公司名称", "价税合计", "货币", "源文件名", "商品列表"
        ]

        # 确保列存在才排序
        existing_cols = [c for c in priority_cols if c in df.columns]
        other_cols = [c for c in df.columns if c not in existing_cols]
        df = df[existing_cols + other_cols]

        # 保存Excel
        output_path = os.path.join(INPUT_FOLDER, OUTPUT_EXCEL)
        df.to_excel(output_path, index=False, engine='openpyxl')

        # 添加超链接到源文件
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        from openpyxl.worksheet.hyperlink import Hyperlink

        wb = load_workbook(output_path)
        ws = wb.active

        # 找到"源文件名"列的索引
        header_row = 1
        source_file_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col).value == "源文件名":
                source_file_col = col
                break

        if source_file_col:
            # 为每个源文件名添加超链接
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=source_file_col)
                source_filename = cell.value

                if source_filename and not pd.isna(source_filename):
                    # 从 "travel_493193.pdf (第1页)" 提取文件名
                    filename = source_filename.split(' (')[0] if ' (' in source_filename else source_filename

                    # 构建完整文件路径
                    full_path = os.path.join(INPUT_FOLDER, filename)

                    # 检查文件是否存在
                    if os.path.exists(full_path):
                        # 转换为文件路径URL格式（Mac）
                        file_url = f"file://{full_path}"
                        cell.hyperlink = Hyperlink(target=file_url, ref=cell.coordinate)
                        cell.style = "Hyperlink"
                        # 保持显示的文本不变
                        cell.value = source_filename

            # 保存修改
            wb.save(output_path)

        print(f"✅ Excel已生成: {output_path}")
        print(f"📄 共 {len(all_results)} 行数据")
        print(f"🔗 源文件名列已添加超链接，点击可直接打开原始文件\n")

        # 显示前5行预览
        print("="*60)
        print("📋 数据预览（前5行）:")
        print("="*60)
        print(df.head().to_string(index=False))
        print("="*60)

    else:
        print("❌ 没有提取到任何数据")


if __name__ == "__main__":
    main()
